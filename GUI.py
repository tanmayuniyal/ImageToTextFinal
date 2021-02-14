#%%
from tkinter import *
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import main as bg

wb = Workbook() 
wb = load_workbook('translate.xlsx')
ws = wb.active

column_a = ws ['A']
column_b = ws ['B']
column_c = ws ['C']
column_d = ws ['D']
column_e = ws ['E']
column_f = ws ['F']
column_g = ws ['G']

def get_a():
    list = ''
    for cell in column_a:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_b():
    list = ''
    for cell in column_b:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_c():
    list = ''
    for cell in column_c:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_d():
    list = ''
    for cell in column_d:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_e():
    list = ''
    for cell in column_e:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_f():
    list = ''
    for cell in column_f:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_g():
    list = ''
    for cell in column_g:
        list = f"{list + str(cell.value)}\n"
    Label(window, text = list ).pack(padx = 20)

def get_file_path():
    global file_path
    # Open and return file path
    file_path= filedialog.askopenfilename(title = "Select A File", filetypes = (("type1", "*.png"), ("type2", "*.jpg")))
    l1 = Label(window, text = "File path: " + file_path).pack()

def submit():
    bg.background(file_path)


window = Tk()
window.title("Image to Text")
window.geometry("500x500")

b1 = Button(window, text = "Open File", command = get_file_path).pack()

b2 = Button(window, text= "Submit", command= submit).pack(padx = 20)

bA = Button(window, text = "Words", command = get_a).pack(padx = 20)

bB = Button(window, text = "Synonym", command = get_b).pack(padx = 20)

bC = Button(window, text = "Antonym", command = get_c).pack(padx = 20)

bD = Button(window, text = "Hindi", command = get_d).pack(padx = 20)

bE = Button(window, text = "Punjabi", command = get_e).pack(padx = 20)

bF = Button(window, text = "Tamil", command = get_f).pack(padx = 20)

bG = Button(window, text = "Telugu", command = get_g).pack(padx = 20)

window.mainloop()


# %%
